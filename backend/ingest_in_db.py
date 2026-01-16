# import basics
import os
from dotenv import load_dotenv
import openpyxl

# import langchain
from langchain_community.document_loaders import PyPDFDirectoryLoader
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain_community.vectorstores import SupabaseVectorStore
 
from langchain_core.documents import Document

# import supabase
from supabase.client import Client, create_client

# load environment variables
load_dotenv()

# initiate supabase db
supabase_url = os.getenv("SUPABASE_URL")
supabase_key = os.getenv("SUPABASE_SERVICE_KEY")
supabase: Client = create_client(supabase_url, supabase_key)

# initiate Google Generative AI embeddings
from langchain_google_genai import GoogleGenerativeAIEmbeddings
GOOGLE_API_KEY = os.environ.get("GOOGLE_API_KEY")
embeddings = GoogleGenerativeAIEmbeddings(
    model="models/embedding-001",
    google_api_key=GOOGLE_API_KEY
)

# load pdf docs from folder 'documents'
loader = PyPDFDirectoryLoader("documents")
documents = loader.load()

# Load Excel files
def load_excel_as_documents(directory):
    excel_docs = []
    for filename in os.listdir(directory):
        if filename.endswith(".xlsx"):
            filepath = os.path.join(directory, filename)
            try:
                # Specific handling for RiceRecommendationData which has messy headers
                if "RiceRecommendationData" in filename:
                    wb = openpyxl.load_workbook(filepath, read_only=True)
                    ws = wb.active
                    rows = list(ws.iter_rows(values_only=True))
                    
                    if not rows:
                        continue
                        
                    # Find header row: look for a row with "Disease Name" or similar known columns
                    # Or heuristic: first row with > 3 non-empty string values
                    header_index = -1
                    headers = []
                    
                    for i, row in enumerate(rows):
                        # Heuristic: check for meaningful columns we saw in analysis
                        row_str = [str(c) for c in row if c is not None]
                        if any("Disease Name" in str(c) for c in row if c) or \
                           any("Management" in str(c) for c in row if c) or \
                           len(row_str) >= 4:
                             header_index = i
                             headers = row
                             break
                    
                    if header_index == -1:
                        print(f"Could not find likely header in {filename}, skipping.")
                        continue
                        
                    for row in rows[header_index+1:]:
                        content_parts = []
                        # Skip empty rows
                        if not any(row): 
                            continue
                            
                        for i, cell in enumerate(row):
                            if cell is not None:
                                header = headers[i] if i < len(headers) and headers[i] else f"Column {i}"
                                # Clean up header and cell
                                header = str(header).strip().replace('\n', ' ')
                                cell_val = str(cell).strip()
                                content_parts.append(f"{header}: {cell_val}")
                        
                        if content_parts:
                            content = "Rice Recommendation Data:\n" + "\n".join(content_parts)
                            excel_docs.append(Document(page_content=content, metadata={"source": filename}))

                else:
                    # Standard ingestion for other files (like Cleaned_Data if we want to ingest it too, 
                    # though plan implies we might just use Cleaned_Data for reports. 
                    # If we ingest Cleaned_Data, it creates huge number of docs. 
                    # Implementation plan didn't explicitly say NOT to ingest Cleaned_Data, 
                    # but current code does. Let's keep it safe and robust.)
                    wb = openpyxl.load_workbook(filepath, read_only=True)
                    ws = wb.active
                    rows = list(ws.iter_rows(values_only=True))
                    if not rows:
                        continue
                    
                    headers = rows[0]
                    for row in rows[1:]:
                        content_parts = []
                        for i, cell in enumerate(row):
                            if cell is not None:
                                header = headers[i] if i < len(headers) else f"Column {i}"
                                content_parts.append(f"{header}: {cell}")
                        
                        if content_parts:
                            content = "\n".join(content_parts)
                            excel_docs.append(Document(page_content=content, metadata={"source": filename}))
                            
            except Exception as e:
                print(f"Error loading {filename}: {e}")
    return excel_docs

excel_documents = load_excel_as_documents("documents")
documents.extend(excel_documents)

# split the documents in multiple chunks
text_splitter = RecursiveCharacterTextSplitter(chunk_size=1000, chunk_overlap=200)
docs = text_splitter.split_documents(documents)

# store chunks in Supabase vector store
print(f"Ingesting {len(docs)} chunks into Supabase...")
vector_store = SupabaseVectorStore.from_documents(
    docs,
    embeddings,
    client=supabase,
    table_name="documents",
    query_name="match_documents",
    chunk_size=500,
)
print("Ingestion complete.")